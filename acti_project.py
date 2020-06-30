from __future__ import print_function
import pickle
import os.path
from googleapiclient.discovery import build
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
import pandas as pd
from openpyxl import load_workbook
import openpyxl
import datetime
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
from googleapiclient.http import MediaFileUpload
import sys
from httplib2 import Http
from json import dumps

script_dir = os.path.dirname(__file__)

# If modifying these scopes, delete the file token.pickle.
SCOPES = ['https://www.googleapis.com/auth/drive']


class GoogleDrive:

    def login(self):
        """
        # USe cookies
            Shows basic usage of the Drive v3 API.
            returns googleapiclient.discovery.Resource object
        """
        creds = None
        if os.path.exists('/var/lib/jenkins/workspace/Actitime-Summary-Report/actiTime-project/token.pickle'):
            with open('/var/lib/jenkins/workspace/Actitime-Summary-Report/actiTime-project/token.pickle', 'rb') as token:
                creds = pickle.load(token)
        # If there are no (valid) credentials available, let the user log in.
        if not creds or not creds.valid:
            if creds and creds.expired and creds.refresh_token:
                creds.refresh(Request())
            else:
                flow = InstalledAppFlow.from_client_secrets_file(
                    '/var/lib/jenkins/workspace/Actitime-Summary-Report/actiTime-project/credentials.json', SCOPES)
                creds = flow.run_local_server(port=0)
            # Save the credentials for the next run
            with open('/var/lib/jenkins/workspace/Actitime-Summary-Report/actiTime-project/token.pickle', 'wb') as token:
                pickle.dump(creds, token)

        service = build('drive', 'v3', credentials=creds)

        return service

    def parent_folder_identifier(self, folder_name):
        """
            :param: folder_name (unique parent folder name to search)
            :return: parent_folder_id
        """
        drive = self.login()
        parent_folder_id = None
        page_token = None
        while True:
            response = drive.files().list(q="mimeType='application/vnd.google-apps.folder'" and "name contains '" + folder_name + "'",
                                            spaces='drive',
                                            fields='nextPageToken, files(id, name)',
                                            pageToken=page_token).execute()
            for file in response.get('files', []):
                if file.get('name').find("pdf") == -1:
                    parent_folder_id = file

            page_token = response.get('nextPageToken', None)
            if page_token is None:
                break

        return parent_folder_id

    def child_folder_identifier(self, parent_folder_id, child_folder_name):
        """
            :param:
                 parent_folder_id : comes from parent_folder_identifier()^^
                 child_folder_name: child folder name from parent folder
            :return: child_folder_id
        """
        drive = self.login()
        child_folder_id = None
        page_token = None
        while True:
            response = drive.files().list(
                        q="mimeType='application/vnd.google-apps.folder'" and "'" + parent_folder_id + "' in parents",
                        spaces='drive',
                        fields='nextPageToken, files(id, name)',
                        pageToken=page_token).execute()
            for file in response.get('files', []):
                if (file.get('name')) == child_folder_name:
                    child_folder_id = file
            page_token = response.get('nextPageToken', None)
            if page_token is None:
                break

        return child_folder_id

    def list_files(self, folder_id, report=None):
        """
            :param
                folder_id: identifier of the folder name.
            :return: lists all files present in the folder.
        """
        drive = self.login()
        files = []
        page_token = None
        while True:
            response = drive.files().list(
                        q="mimeType='application/vnd.google-apps.file'" and "'" + folder_id + "' in parents",
                        spaces='drive',
                        fields='nextPageToken, files(id, name)',
                        pageToken=page_token).execute()
            for file in response.get('files', []):
                if report == "workbook":
                    if file.get('name').find("OPP") == -1:
                        pass
                    else:
                        files.append(file)
                if report == "actiTime":
                    if file.get('name').find("weekending") == -1:
                        pass
                    else:
                        files.append(file)

            page_token = response.get('nextPageToken', None)
            if page_token is None:
                break

        return files

    def download_files(self, file_ids, report=None):

        drive = self.login()

        if file_ids:
            request = drive.files().get_media(fileId=file_ids['id'])
            result = request.execute()
            if report == "workbook":
                rel_path = "workbooks/{}".format(file_ids['name'])
                path = os.path.join(script_dir, rel_path)

                print("Downloading file {}".format(file_ids['name']))
                with open(path, 'wb') as f:
                    f.write(result)
            if report == "actiTime":
                rel_path = "reports/{}".format(file_ids['name'])
                path = os.path.join(script_dir, rel_path)

                print("Downloading file {}".format(file_ids['name']))
                with open(path, 'wb') as f:
                    f.write(result)

    def upload_file(self):
        drive = self.login()
        # file_metadata = {'name': 'Actitime-Summary-Copy.xlsx', 'mimeType': 'application/vnd.google-apps.spreadsheet',
        #                  'parents': ['1MjXZyeB_y0C32St20-yXPWDA-H4pcAau']}
        file_metadata = {'name': 'Actitime-Summary-Copy.xlsx', 'mimeType': 'application/vnd.google-apps.spreadsheet'}

        media = MediaFileUpload('/var/lib/jenkins/workspace/Actitime-Summary-Report/actiTime-project/reports/Actitime-Summary.xlsx', resumable=True,
                                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        file = drive.files().create(body=file_metadata,
                                    media_body=media).execute()

        # print("Created file '%s' id '%s'." % (file.get('name'), file.get('id')))

        return file

    def fetch_acti_data(self, week_day, acti_summary_file, custom_file):
        xls = pd.ExcelFile(acti_summary_file)
        df1 = pd.read_excel(xls, 'projects-detail')

        search = week_day
        index_val = [(df1[col][df1[col].eq(search)].index[i],
                      df1.columns.get_loc(col)) for col in df1.columns for i in
                     range(len(df1[col][df1[col].eq(search)].index))]

        col_index = index_val[0][1]

        df = pd.read_excel(acti_summary_file, sheet_name='projects-detail', usecols=[0, col_index])

        df.columns = ['Projects', 'ActiHours']
        df1 = df.iloc[4:]

        opp_id = []
        project_names = []
        acti_hours = []

        for proj, actihrs in zip(df1['Projects'], df1['ActiHours']):
            opp_id.append("".join(proj.split())[:9])
            project_names.append(proj)
            acti_hours.append(actihrs)

        df1.insert(0, "Opp-ID", opp_id, True)

        writer = pd.ExcelWriter(custom_file, engine='openpyxl')
        df1.to_excel(writer, sheet_name='actiSource', index=False)
        writer.save()

        return opp_id, project_names, acti_hours, col_index

    def insert_wkb_hours(self, opp_file, custom_file, day, year, actiData):
        df = pd.read_csv(opp_file)
        search_val = df.loc[:,df.columns.str.contains(str(day))].fillna(0)

        total_workbook_hrs = 0
        for col in search_val.columns:
            if str(search_val[col].iloc[0]).split('.')[0] == year:
                search_val1 = search_val[col].values.tolist()
                del search_val1[0:2]
                total_workbook_hrs = sum(search_val1) * 8
        for i in actiData:
            if opp_file.strip("/var/lib/jenkins/workspace/Actitime-Summary-Report/actiTime-project/workbooks/OPP-").strip(".csv") in i:
                index_value = actiData.index(i) + 1
                df2 = pd.DataFrame({total_workbook_hrs})
                print("{}  => workbookhours data => {}".format(i, total_workbook_hrs))
                fn = custom_file
                writer = pd.ExcelWriter(fn, engine='openpyxl')
                book = load_workbook(fn)
                writer.book = book
                writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                df2.to_excel(writer, sheet_name='actiSource', header=None, index=False, startcol=3, startrow=index_value)
                writer.save()

    def insert_difference(self, custom_file):
        df = pd.read_excel(custom_file, sheet_name='actiSource', usecols="C:D")
        df.columns = ['acti_hours', 'workbook_hours']
        df_wkb_hours = df['workbook_hours'].fillna(0)
        difference = []
        for act, wkb in zip(df['acti_hours'], df_wkb_hours):
            hours_diff = wkb - act
            difference.append(hours_diff)

        wb = openpyxl.load_workbook(custom_file)
        ws = wb['actiSource']

        for row, v in enumerate(difference, 2):
            ws.cell(row, 5, v)

        # save the changes to the workbook
        wb.save(custom_file)

    def generate_report(self, acti_summary_file, custom_file, day):
        xls = pd.ExcelFile(acti_summary_file)
        acti_df = pd.read_excel(xls, 'projects-detail')

        ds_df = pd.read_excel(custom_file, sheet_name='actiSource', usecols="D:E")
        ds_df.columns = ['wkb_hours', 'diff']

        diff_hours = ds_df['diff'].values.tolist()
        wkb = ds_df['wkb_hours'].values.tolist()

        wkb_hours = []
        for hrs in range(len(wkb)):
            if str(wkb[hrs]) == 'nan':
                wkb_hours.append(str(wkb[hrs]).replace('nan', 'No WKB data'))
            elif str(wkb[hrs]) == "0.0":
                wkb_hours.append(str(wkb[hrs]).replace('0.0', 'No WKB data'))
            else:
                wkb_hours.append(wkb[hrs])

        search = day
        index_val = [(acti_df[col][acti_df[col].eq(search)].index[i],
                      acti_df.columns.get_loc(col)) for col in acti_df.columns for i in
                     range(len(acti_df[col][acti_df[col].eq(search)].index))]

        col_index = index_val[0][1] + 1
        wkb_col_index = col_index + 1
        diff_col_index = col_index + 2

        wb = openpyxl.load_workbook(acti_summary_file)
        ws = wb['projects-detail']

        for wkb_hrs, wkb_cell in enumerate(wkb_hours, 6):
            ws.cell(wkb_hrs, wkb_col_index, wkb_cell)

        for diff_hrs, diff_cell in enumerate(diff_hours, 6):
            ws.cell(diff_hrs, diff_col_index, diff_cell)

        count = 6
        for i in range(len(wkb_hours)):
            ws[get_column_letter(wkb_col_index) + str(count)].font = Font(name='Calibri', bold=False, size=10)
            ws[get_column_letter(wkb_col_index) + str(count)].alignment = Alignment(horizontal='center')

            ws[get_column_letter(diff_col_index) + str(count)].font = Font(name='Calibri', bold=False, size=10)
            ws[get_column_letter(diff_col_index) + str(count)].alignment = Alignment(horizontal='center')

            if str(ws[get_column_letter(wkb_col_index) + str(count)].value) == "No WKB data":
                ws[get_column_letter(wkb_col_index) + str(count)].font = Font(name='Calibri', color="FF0000",
                                                                              bold=False, size=10)

            if ws[str(get_column_letter(diff_col_index)) + str(count)].value < 0:
                ws[get_column_letter(diff_col_index) + str(count)].fill = PatternFill("solid", fgColor="D23F2A")
                ws[get_column_letter(diff_col_index) + str(count)].font = Font(name='Calibri', color="FFFFFF",
                                                                               bold=False, size=10)

            count += 1

        wb.save(acti_summary_file)

    def smartshift_Bot(self, message, webhook_url):

        url = webhook_url
        bot_message = {
                    "text": message
        }

        message_headers = {'Content-Type': 'application/json; charset=UTF-8'}

        http_obj = Http()

        response = http_obj.request(
                uri=url,
                method='POST',
                headers=message_headers,
                body=dumps(bot_message),
        )

        return response


def main(args):
    GDrive = GoogleDrive()
    # all_workbooks = GDrive.list_files('1A9Aak_IosxU1ylmflkC-s_Mw65GvM7AV', report="workbook")
    all_workbooks = os.listdir('/var/lib/jenkins/workspace/Actitime-Summary-Report/actiTime-project/workbooks/')

    custom_file = '/var/lib/jenkins/workspace/Actitime-Summary-Report/actiTime-project/reports/dataframe.xlsx'
    acti_summary_file = '/var/lib/jenkins/workspace/Actitime-Summary-Report/actiTime-project/reports/Actitime-Summary.xlsx'

    acti_cal_day = datetime.datetime.strptime(args[1], '%d-%m-%Y')
    year = acti_cal_day.strftime("%Y")

    next_day = acti_cal_day + datetime.timedelta(days=1)

    date_val = next_day.strftime("%d")
    month_name = next_day.strftime("%b")

    wkb_cal_day = date_val + '-' + month_name

    actiData = GDrive.fetch_acti_data(acti_cal_day, acti_summary_file, custom_file)

    for wkbs in all_workbooks:
       GDrive.insert_wkb_hours('/var/lib/jenkins/workspace/Actitime-Summary-Report/actiTime-project/workbooks/{}'.format(wkbs), custom_file, wkb_cal_day, year, actiData[0])

    GDrive.insert_difference(custom_file)

    GDrive.generate_report(acti_summary_file, custom_file, acti_cal_day)

    if args[2] == "Yes":
        file = GDrive.upload_file()
        url = "https://chat.googleapis.com/v1/spaces/34csxgAAAAE/messages?key=AIzaSyDdI0hCZtE6vySjMm-WEfRq3CPzqKqqsHI&token=bGngV0oepF6rKb2tayKeAZ4p0Gix9ahwk7yyeHbhb20%3D"
        message = "*{}: {}.xlsx GDrive Link ->* <https://docs.google.com/spreadsheets/d/{}>\n\n Testing..".format(acti_cal_day.date(), file.get('name'), file.get('id'))
        GDrive.smartshift_Bot(message, url)
    else:
        pass


if __name__ == '__main__':
    main(sys.argv)

